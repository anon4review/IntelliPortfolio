import sqlite3
import openpyxl
lists = sqlite3.connect('G:\组合投资\PGPortfolio-master\database\DataStockLong.db')
c = lists.cursor()
listinsheet = openpyxl.load_workbook(r'C:\Users\user\Desktop\数据\上证50长期.xlsx')
datainlist = listinsheet.active  # 获取excel文件当前表格
data_truck='''INSERT INTO main.History(date,coin,high,low,open,close,volume,quoteVolume,weightAverage) VALUES (?,?,?,?,?,?,?,?,?)'''
for row in datainlist.iter_rows(min_row=2, max_col=9, max_row=datainlist.max_row):  # 使excel各行数据成为迭代器
    cargo = [cell.value for cell in row]  # 使每行中单元格成为迭代器
    c.execute(data_truck, cargo)  # 写入一行数据到数据库中表History
lists.commit()
lists.close()

# listinsheet = openpyxl.load_workbook(r'C:\Users\user\Desktop\数据\上证50指数点位.xlsx')
# datainlist = listinsheet.active  # 获取excel文件当前表格
# data_truck='''INSERT INTO main.stock_index(date,point) VALUES (?,?)'''
# for row in datainlist.iter_rows(min_row=2, max_col=2, max_row=datainlist.max_row):  # 使excel各行数据成为迭代器
#     cargo = [cell.value for cell in row]  # 使每行中单元格成为迭代器
#     c.execute(data_truck, cargo)  # 写入一行数据到数据库中表Index
# lists.commit()
# lists.close()