#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
import os
import collections
import tflearn
import numpy as np
import tensorflow as tf
from pgportfolio.learn.nnagent import NNAgent
from pgportfolio.marketdata.datamatrices import DataMatrices
import logging
from openpyxl.workbook import Workbook
Result = collections.namedtuple("Result",
                                [
                                 "test_pv",
                                 "test_log_mean",
                                 "test_log_mean_free",
                                 "test_history",
                                 "config",
                                 "net_dir",
                                 "backtest_test_pv",
                                 "backtest_test_history",
                                 "backtest_test_log_mean",
                                 "training_time"])

class TraderTrainer:
    def __init__(self, config, device="cpu"):
        """
        :param config: config dictionary
        :param device: the device used to train the network
        """
        self.config = config
        self.train_config = config["training"]
        self.input_config = config["input"]
        self.best_metric = 0
        # np.random.seed(config["random_seed"])

        self.__window_size = self.input_config["window_size"]
        self.__coin_number = self.input_config["coin_number"]
        self.__batch_size = self.train_config["batch_size"]

        self._matrix = DataMatrices.create_from_config(config)
        self.stock_code = self._matrix.stock_code

        self.test_set = self._matrix.get_test_set()
        # if not config["training"]["fast_train"]:
        self.training_set = self._matrix.get_training_set()
        self.upperbound_validation = 1
        self.upperbound_test = 1
        # tf.set_random_seed(self.config["random_seed"])
        self.device = device
        if device == "cpu":
            os.environ["CUDA_VISIBLE_DEVICES"] = ""
            with tf.device("/cpu:0"):
                self._agent = NNAgent(config, device)
        else:
            self._agent = NNAgent(config, device)

    def _evaluate(self, set_name, *tensors):
        if set_name == "test":
            feed = self.test_set
        elif set_name == "training":
            feed = self.training_set
        else:
            raise ValueError()
        result,output = self._agent.evaluate_tensors(feed["X"],feed["y"],last_w=feed["last_w"],
                                                     setw=feed["setw"], tensors=tensors, stock_index=feed["stock_index"],
                                                     # market_capticalization=feed["market_capticalization"],
                                                     # all_market_capticalization=feed["all_market_capticalization"]
                                                     )

        if set_name == "test":
            res = []
            for i in range(self.__coin_number):
                rowtemp = []
                rowtemp.append(self.stock_code[i])
                for j in range(60):
                    if j < self.__window_size - 1:
                        rowtemp.append(0.0)
                    elif j == self.__window_size - 1:
                        rowtemp.append(1.0/self.__coin_number)
                    else:
                        rowtemp.append(output[j - self.__window_size][i])
                res.append(rowtemp)
            print(res)
            outwb = Workbook()
            wo = outwb.active
            sheet = outwb.create_sheet('decision', 0)

            sheet.cell(1, 1).value = ''
            for col in range(60):
                sheet.cell(1, col + 2).value = col + 1
            for row in range(self.__coin_number):
                for col in range(61):
                    sheet.cell(row + 2, col + 1).value = res[row][col]
            outwb.save('F:\portfolio_rx\portfolio_rx\\N225\\vol_pg_result.xlsx')
        return result

    @staticmethod
    def calculate_upperbound(y):
        array = np.maximum.reduce(y[:, 0, :], 1)
        total = 1.0
        for i in array:
            total = total * i
        return total

    def log_between_steps(self, step):
        tflearn.is_training(False, self._agent.session)

        # v_pv, v_log_mean, v_loss, log_mean_free, weights, tracking_error, excess_return, sharpe_ratio, information_ratio, tracking_ratio= \
        v_pv, v_log_mean, v_loss, log_mean_free, weights, tracking_error, excess_return, sharpe_ratio, information_ratio =\
            self._evaluate("training",
                           self._agent.portfolio_value,
                           self._agent.log_mean,
                           self._agent.loss,
                           self._agent.log_mean_free,
                           self._agent.portfolio_weights,
                           self._agent.tracking_error,
                           self._agent.excess_return,
                           self._agent.sharp_ratio,
                           self._agent.information_ratio,
                           # self._agent.tracking_ratio
                           )

        loss_value = self._evaluate("training",self._agent.loss)

        print('='*30)
        print('step %d' % step)
        print('-'*30)
        print('the portfolio value on training set is %s\nlog_mean is %s\n'
                     'loss_value is %3f\nlog mean without commission fee is %3f\ntracking error is %3f\n'
                     'excess_return is %3f\nsharpe_ratio is %3f\ninformation_ratio is %3f\n'% \
                     (v_pv, v_log_mean, v_loss, log_mean_free, tracking_error, excess_return, sharpe_ratio, information_ratio))
        # print('tracking_ratio is '+str(tracking_ratio))
        print('='*30+"\n")


    def next_batch(self):
        batch = self._matrix.next_batch()
        batch_input = batch["X"]
        batch_y = batch["y"]
        batch_last_w = batch["last_w"]
        batch_w = batch["setw"]
        batch_stock_index = batch["stock_index"]
        # batch_market_capticalization = batch["market_capticalization"]
        # batch_all_market_capticalization = batch["all_market_capticalization"]
        # return batch_input, batch_y, batch_last_w, batch_w, batch_stock_index, batch_market_capticalization,batch_all_market_capticalization
        return batch_input, batch_y, batch_last_w, batch_w, batch_stock_index

    def __print_upperbound(self):
        upperbound_test = self.calculate_upperbound(self.test_set["y"])
        logging.info("upper bound in test is %s" % upperbound_test)

    def train_net(self):
        self.__print_upperbound()
        for i in range(self.train_config["steps"]):
            # x, y, last_w, setw, stock_index, market_capticalization, all_market_capticalization = self.next_batch()
            x, y, last_w, setw, stock_index = self.next_batch()
            # self._agent.train(x, y, last_w=last_w, setw=setw, stock_index=stock_index, market_capticalization=market_capticalization, all_market_capticalization=all_market_capticalization)
            self._agent.train(x, y, last_w=last_w, setw=setw, stock_index=stock_index)
            if i % 1000 == 0:
                self.log_between_steps(i)

        sr, te,er,ir = self._evaluate("test", self._agent.sharp_ratio, self._agent.tracking_error,self._agent.excess_return,self._agent.information_ratio)
        print('test set:sharp ratio is %lf, TE is %lf, ER is %lf, IR is %lf'% ( sr, te,er,ir))
