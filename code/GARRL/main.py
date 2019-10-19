from GARRL.garrl import GARRL
import numpy as np
import tensorflow as tf
import configparser
import tqdm
from openpyxl.workbook import Workbook

def getConfig(section, key):
    config = configparser.ConfigParser()
    path = './config.ini'
    config.read(path,encoding='utf-8')
    return config.get(section, key)

def init_avg_elitist(avg_eli,list_choosen_elis,sess,saver):
    saver.restore(sess, tf.train.latest_checkpoint('./'))
    var_dict = {}
    for i,eli in enumerate(list_choosen_elis):
        vars=sess.run(eli.vars)
        if i==0:
            index=0
            for j in vars:
                k = avg_eli.var_holder[index]
                var_dict[k] = j
                index+=1
        else:
            index = 0
            for j in vars:
                k = avg_eli.var_holder[index]
                var_dict[k] +=j
                index+=1
    for holder in var_dict:
        var_dict[holder] /= len(list_choosen_elis)
    sess.run(avg_eli.assign_step,feed_dict=var_dict)


if __name__ =='__main__':
    choosen_stocks_num=int(getConfig('rrl', 'choosen_stocks_num'))
    win_len=int(getConfig('rrl', 'window_len'))
    garrl = GARRL()
    garrl.ga()
    best_DSR=garrl.best_fitness
    indicator_type=garrl.best_indicator_type
    print('best indicators type choosen complete!')
    print('==============================================')
    datag=garrl.elitistsSystem.datag
    list_choosen_elis=garrl.best_choosen_elistists
    avg_eli=garrl.avg_elitist

    init_avg_elitist(avg_eli,list_choosen_elis,garrl.sess,garrl.saver)
    print('avg eli init complete! begin trade training ')
    print('==============================================')

    #trade train_loop
    trade_data = datag.get_trade_data(indicator_type)
    trade_index=datag.trade_index
    trade_num = int(getConfig('trade', 'trade_num'))
    loop_time=int(getConfig('train', 'trade_train_epoch'))
    one_step = int(getConfig('rrl', 'one_step'))
    for step in tqdm.tqdm(range(loop_time)):
        _current_cell_state_avg = np.zeros((1, choosen_stocks_num))
        _current_hidden_state_avg = np.array([datag.Ftminus1])
        Ftminus1_avg = datag.Ftminus1
        for y in range(int((trade_num-win_len+1) / one_step)):
            feed = {avg_eli.input: trade_data[:,y*one_step:one_step+one_step*y,:],
                    avg_eli.Ftminus1:Ftminus1_avg ,
                    avg_eli.price: datag.trade_price[:,y*one_step:one_step+one_step*y],
                    avg_eli.cell_state: _current_cell_state_avg,
                    avg_eli.hidden_state: _current_hidden_state_avg,
                    avg_eli.index:trade_index[y*one_step:one_step+one_step*y]
                    }
            _,state,Ft=garrl.sess.run([avg_eli.train_step_noholder,avg_eli.state,avg_eli.origin_outputs],feed_dict=feed)
            _current_cell_state_avg, _current_hidden_state_avg = state
            Ftminus1_avg=Ft[-1]
        _current_cell_state_avg = np.zeros((1, choosen_stocks_num))
        _current_hidden_state_avg =np.array([datag.Ftminus1])
        Ftminus1_avg = datag.Ftminus1
        feed2={avg_eli.input: trade_data,
                    avg_eli.Ftminus1: Ftminus1_avg,
                    avg_eli.price: datag.trade_price,
                    avg_eli.cell_state: _current_cell_state_avg,
                    avg_eli.hidden_state: _current_hidden_state_avg,
                    avg_eli.index:trade_index
                    }
        sr, ER,TE = garrl.sess.run([avg_eli.sharpe_ratio, avg_eli.ER,avg_eli.TE], feed_dict=feed2)
        print('sharpe_ratio',sr)
        print('ER', ER)
        print('TE', TE)

    print('trade training complete! begin test ')
    print('==============================================')
    # test
    _current_cell_state = np.zeros((1, choosen_stocks_num))
    _current_hidden_state =np.array([datag.Ftminus1])
    test_data = datag.get_test_data(indicator_type)
    feed = {avg_eli.input: test_data,
            avg_eli.Ftminus1: datag.Ftminus1,
            avg_eli.price: datag.test_price,
            avg_eli.index:datag.test_index,
            avg_eli.cell_state: _current_cell_state,
            avg_eli.hidden_state: _current_hidden_state
            }
    SR, Ft,ER,TE = garrl.sess.run([avg_eli.sharpe_ratio, avg_eli.origin_outputs,avg_eli.ER,avg_eli.TE], feed_dict=feed)
    print('test_SR:', SR)
    print('decision:', Ft[-1])
    print('ER:',ER)
    print('TE',TE)

    # result.xlsx
    choosen_stocks_codes=datag.code
    result=[]
    for i in range(choosen_stocks_num):
        rowtemp=[]
        rowtemp.append(choosen_stocks_codes[i])
        for j in range(60):
            if j<win_len:
                rowtemp.append(0.1)
            else:
                rowtemp.append(Ft[j-win_len][i])
        result.append(rowtemp)

    outwb = Workbook()
    wo = outwb.active
    sheet = outwb.create_sheet('decision',0)

    sheet.cell(1, 1).value=''
    for col in range(60):
        sheet.cell(1, col + 2).value = col+1
    for row in range(choosen_stocks_num):
        for col in range(61):
            sheet.cell(row+2,col+1).value=result[row][col]
    outwb.save('..\FTSE\\vol_garrl_result.xlsx')