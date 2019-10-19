from heuristicGA.ga import GA
from heuristicGA.TLtest import TLTest
import numpy as np
import openpyxl
from openpyxl.workbook import Workbook
if __name__=='__main__':
    ga=GA()
    names=ga.ga_loop()

    decisions=ga.decision

    test=TLTest()
    decisions=np.array(decisions)
    # 最后一天decisions无用
    test_object=test.get_L_object(decisions[0:-1])
    decison2excel=decisions[0:-1] #除去最后一个决策
    codes=ga.datakeeper.code

    outwb = Workbook()
    wo = outwb.active
    sheet = outwb.create_sheet('decision', 0)

    sheet.cell(1, 1).value = ''
    for col in range(60):
        sheet.cell(1, col + 2).value = col + 1
    for row in range(ga.N):
        sheet.cell(row + 2, 1).value = codes[row]
        for col in range(60):
            sheet.cell(row + 2, col + 2).value = decison2excel[col][row]
    outwb.save('..\\N225\heuristicGA.xlsx')

    print('Share of shares:',decisions)
    print('T tO T+L-1 tracing_error:',test_object[0])
    print('T tO T+L-1 excess_return:', test_object[1])
    print('T tO T+L-1 sharpe_ratio:', test_object[2])
    print('T tO T+L-1 IR:', test_object[3])
    # print('T tO T+L-1 tracing_ratio:', test_object[4])

