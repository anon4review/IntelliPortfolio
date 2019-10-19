import configparser
import os
from rrl_lstm import RRL
from getData import GetData
import tensorflow as tf
import tqdm
import openpyxl
from openpyxl.workbook import Workbook

def getConfig(section, key):
    config = configparser.ConfigParser()
    path = os.path.split(os.path.realpath(__file__))[0]+'/config.ini'
    config.read(path)
    return config.get(section, key)

if __name__ == '__main__':
    choosen_stocks_num = int(getConfig('rrl', 'choosen_stocks_num'))
    window_len = int(getConfig('rrl', 'window_len'))
    rrl = RRL()
    getdata = GetData()
    print('train begin')
    print("============================================")
    train_data = getdata.train1()
    feed = {rrl.input: train_data,
            rrl.Ftminus1: getdata.Ftminus1,
            # rrl.cash: [[getdata.cash]],
            rrl.price: getdata.train_price}
    with tf.Session() as sess:
        sess.run(tf.initialize_all_variables())
        for i in tqdm.tqdm (range(100)):
            print('\n')
            _, sr, ft = sess.run([rrl.train_step_noholder,rrl.sharpe_ratio, rrl.outputs],feed_dict=feed)
            print('sharpe_ratio:', sr)
            print('Ft:', ft[-1])

        print('\ntest begin')
        print("=============================================")
        test_data = getdata.test()
        feed = {rrl.input: test_data,
                rrl.Ftminus1: getdata.Ftminus1,
                # rrl.cash: [[getdata.cash]],
                rrl.price: getdata.test_price,
                rrl.index: getdata.test_index,
                rrl.market_capitalization: getdata.test_market_capitalization,
                rrl.all_market_capitalization: getdata.test_all_market_capitalization
                }

        SR, FT, ER, TE, IR, tracking_ratio = sess.run([rrl.sharpe_ratio, rrl.origin_outputs, rrl.ER, rrl.TE, rrl.IR, rrl.tracking_ratio], feed_dict=feed)
        print('test_sharpe_ratio:', SR)
        print('test_Ft:', FT[-1])
        print('Excess Return:', ER)
        print('Tracking Error:', TE)
        print('Information Ratio:', IR)
        print('Tracking Ratio:', tracking_ratio)

    choosen_stocks_codes = getdata.stock
    result = []
    for i in range(choosen_stocks_num):
        rowtemp = []
        rowtemp.append(choosen_stocks_codes[i])
        for j in range(60):
            if j < window_len - 1:
                rowtemp.append(0.0)
            elif j == window_len - 1:
                rowtemp.append(1/choosen_stocks_num)
            else:
                rowtemp.append(FT[j - window_len][i])
        result.append(rowtemp)

    outwb = Workbook()
    wo = outwb.active
    sheet = outwb.create_sheet('decision', 0)

    sheet.cell(1, 1).value = ''
    for col in range(60):
        sheet.cell(1, col + 2).value = col + 1
    for row in range(choosen_stocks_num):
        for col in range(61):
            sheet.cell(row + 2, col + 1).value = result[row][col]
    outwb.save('F:\portfolio_rx\portfolio_rx\DJ30\\vol_rrl_result.xls')