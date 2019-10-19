import sqlite3
import openpyxl
import configparser
import os

def getConfig(section, key):
    config = configparser.ConfigParser()
    path = os.path.split(os.path.realpath(__file__))[0] + '/config.ini'
    config.read(path)
    return config.get(section, key)

db_path=getConfig('db', 'db_path')
conn = sqlite3.connect(database=db_path)

SQL='select distinct date from originData'
date_list=conn.execute(SQL).fetchall()
listinsheet = openpyxl.load_workbook('FTSE100index.xlsx')
datainlist = listinsheet.active # 获取excel文件当前表格
last_closeprice=0.0
row=2
index=0
print(date_list.__len__())
while(index<date_list.__len__()):
    print(index)
    dateinexcel=datainlist.cell(row,1).value
    closeprice=datainlist.cell(row,2).value
    if dateinexcel is None or date_list[index][0]<dateinexcel :
        SQL='insert into main.indexes(date, closeprice) VALUES (%d,%lf)'%(date_list[index][0],last_closeprice)
        conn.execute(SQL)
        index+=1
    elif date_list[index][0]==dateinexcel:
        if(closeprice is None):
            closeprice=last_closeprice
        SQL = 'insert into main.indexes(date, closeprice) VALUES (%d,%lf)' % (date_list[index][0],closeprice)
        conn.execute(SQL)
        last_closeprice=closeprice
        row += 1
        index+=1
    else:
        row+=1

conn.commit()
conn.close()